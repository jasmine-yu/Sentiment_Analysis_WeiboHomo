# 随机按行打乱Excel表格
from openpyxl import Workbook
from openpyxl import load_workbook
import random

# 读取文件
wb = load_workbook('D:/weibo/data.xlsx')
# 显示工作表数量
print("---该文件共" + str(len(wb.sheetnames)) + "个工作表---")


def main():
    # 打乱所有工作簿
    for ws in wb:
        print("\n---开始打乱工作表[" + str(ws.title) + "]---")
        # 显示行数
        print("---行数:" + str(ws.max_row) + '---')

        random_data(ws)

    # 文件另存为data_random.xlsx
    wb.save('D:/weibo/data_random.xlsx')
    print("\n---文件保存成功---")
    word = input("\n按下任意键退出")


def random_data(ws):
    """随机打乱数据"""
    print("---正在打乱数据---")
    numbers = list(range(1, ws.max_row + 1))

    # random()函数采用梅森旋转算法生成伪随机数
    random.shuffle(numbers)
    for i in numbers:
        row = ws[i]
        r = []
        for cell in row:
            r.append(cell.value)
        ws.append(r)

    ws.delete_rows(1, ws.max_row // 2)


if __name__ == "__main__":
    main()